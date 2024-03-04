from selenium import webdriver
from xlwt import Workbook
import openpyxl
from openpyxl.styles import PatternFill
import os.path
import time
import os
import detect_01
from selenium.webdriver.common.by import By
from datetime import datetime
import tkinter as tk
import threading
import joblib
import numpy as np
import pandas as pd
import sys



if (os.path.exists('./ComDataSave.xlsx') == False):
    wb = openpyxl.Workbook()
    s1 = wb.worksheets[0]
    s1.title = 'Goods'
    s2 = wb.create_sheet('Category')
    s3 = wb.create_sheet('Search')
    rowNum = [s1.max_row, s2.max_row, s3.max_row]
    color = ['ffffff', '000000']

    for i in range(rowNum[0]):
        fille = PatternFill('solid', fgColor=color[0])
        s1.cell(row=i + 1, column=1, value='').fill = fille
    for i in range(rowNum[1]):
        fille = PatternFill('solid', fgColor=color[0])
        s2.cell(row=i + 1, column=1, value='').fill = fille
    for i in range(rowNum[2]):
        fille = PatternFill('solid', fgColor=color[0])
        s3.cell(row=i + 1, column=1, value='').fill = fille

    s1['B1'] = '年齡'
    s1['C1'] = '性別'
    s1['D1'] = '表情'
    s1['F1'] = 'productName'
    s1['G1'] = 'productURL'
    s1['H1'] = 'tempTcode'
    s1['I1'] = 'time(s)'

    s2['B1'] = '年齡'
    s2['C1'] = '性別'
    s2['D1'] = '表情'
    s2['F1'] = '類別名稱'
    s2['G1'] = '類別網址'
    s2['H1'] = 'tempTcode'

    s3['B1'] = '年齡'
    s3['C1'] = '性別'
    s3['D1'] = '表情'
    s3['F1'] = '查詢關鍵字'
    s3['G1'] = '關鍵字網址'

    wb.save("ComDataSave.xlsx")
else:
    wb = openpyxl.load_workbook('ComDataSave.xlsx')
    s1 = wb['Goods']
    s2 = wb['Category']
    s3 = wb['Search']

# 變數區
detect = False
current_url = ''
current_title = ''
options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)
result = []
global recommand


def watch_url_and_title():
    driver = webdriver.Chrome()
    input_string = sys.stdin.read()
    if(str(input_string) == '0'):
        driver.get('https://www.momoshop.com.tw/main/Main.jsp')
    elif(str(input_string) == '1'):
        driver.get('https://www.ruten.com.tw/')
    global current_url, current_title, number
    global original_window
    global detect
    number = 1

    old_hour = 0
    old_min = 0
    old_sec = 0

    while True:
        original_window = driver.current_window_handle
        for window_handle in driver.window_handles:
            if window_handle != original_window:
                driver.close()
                driver.switch_to.window(window_handle)
                break
        new_url = driver.current_url
        new_title = driver.title
        try:
            if("momoshop" in new_url):
                momo_element = driver.find_element(By.ID, 'tempTCode')
                momo_IDvalue = momo_element.get_attribute("value")
                if (momo_IDvalue == '1999900000' or momo_IDvalue == '4299900000' or momo_IDvalue == '4399900000'):
                    temp = '0'
                elif (momo_IDvalue == '2999900000'):
                    temp = '1'
                elif (momo_IDvalue == '1199900000' or momo_IDvalue == '4599900000'):
                    temp = '2'
                elif (momo_IDvalue == '1299900000' or momo_IDvalue == '2099900000' or momo_IDvalue == '4499900000'):
                    temp = '3'
                elif (momo_IDvalue == '1399900000' or momo_IDvalue == '1499900000'):
                    temp = '4'
                elif (momo_IDvalue == '3199900000' or momo_IDvalue == '1599900000'):
                    temp = '5'
                elif (momo_IDvalue == '2799900000'):
                    temp = '6'
                elif (momo_IDvalue == '4099900000' or momo_IDvalue == '5199900000' or momo_IDvalue == '5399900000'):
                    temp = '7'
                elif (momo_IDvalue == '1899900000' or momo_IDvalue == '4899900000' or momo_IDvalue == '2899900000' or momo_IDvalue == '5099900000'):
                    temp = '8'
                elif (momo_IDvalue == '3999900000' or momo_IDvalue == '1799900000' or momo_IDvalue == '4799900000' or momo_IDvalue == '5299900000'):
                    temp = '9'
                elif (momo_IDvalue == '1699900000' or momo_IDvalue == '4199900000' or momo_IDvalue == '2599900000' or momo_IDvalue == '3899900000'):
                    temp = '10'
                new_IDvalue = temp
            elif("ruten" in new_url):
                ruten_element = driver.find_element(By.ID, 'searchTargetCategory')
                ruten_IDvalue = ruten_element.get_attribute("value")[:4]
                if (ruten_IDvalue == '0011' or ruten_IDvalue == '0021' or ruten_IDvalue == '0008' or ruten_IDvalue == '0022'):
                    temp = '0'
                elif (ruten_IDvalue == '0023'):
                    temp = '1'
                elif (ruten_IDvalue == '0012'):
                    temp = '2'
                elif (ruten_IDvalue == '0024'):
                    temp = '3'
                elif (ruten_IDvalue == '0018' or ruten_IDvalue == '0002'  or ruten_IDvalue == '0014'):
                    temp = '4'
                elif (ruten_IDvalue == '0015'):
                    temp = '5'
                elif (ruten_IDvalue == '0001' or ruten_IDvalue == '0005'):
                    temp = '6'
                elif (ruten_IDvalue == '0006' or ruten_IDvalue == '0004' or ruten_IDvalue == '0013' or ruten_IDvalue == '0010'):
                    temp = '7'
                elif (ruten_IDvalue == '0003' or ruten_IDvalue == '0017'):
                    temp = '8'
                elif (ruten_IDvalue == '0009'):
                    temp = '9'
                elif (ruten_IDvalue == '0020' or ruten_IDvalue == '0019'):
                    temp = '10'
                new_IDvalue = temp
        except:
            new_IDvalue = "7878"
        if new_url != current_url:

            #算時間用
            time = datetime.now()
            min = time.minute
            sec = time.second
            hour = time.hour
            # print(hour,min, sec)
            if(old_hour != None):
                time1 = (((hour-old_hour+24)%24)*60*60) + (min-old_min)*60 + (sec-old_sec)
            #     print('時間:',time1)
            # print(old_hour, old_min, old_sec)
            old_min = min
            old_sec = sec
            old_hour = hour

            gender, age, emotion = detect_01.detection()
            current_url = new_url
            current_title = new_title

            test_data = [[int(gender), int(age), int(emotion), int(time1)]]
            if(new_IDvalue != '7878'):
                knn = joblib.load('knn.model')
                return_type = knn.predict(test_data)
                return_type = np.array_str(return_type)
                tempT1 = []
                tempT2 = []
                df = pd.read_excel('ComDataSave.xlsx')
                y = df['tempTcode'].values
                for i in range(0, len(y)):
                    if str([y[i]]) == return_type:
                        tempT1.append(df.at[i, 'productURL'])
                        tempT2.append(df.at[i, 'productName'])
                global recommand
                index = random.choice([0,len(tempT1)-1])
                recommand = tempT2[index]

                t = threading.Thread(target=monitor_browser_time, args=(str(recommand),return_type))
                t.start()
            judgeWhereToWrite(gender, age, emotion, new_url, new_title,new_IDvalue, time1)
            #print(f"New URL: {new_url}, Title: {new_title}")


def judgeWhereToWrite(gender, age, emotion, url, title, IDvalue,time):
    record = 0
    if ("momoshop" in url):  # momo's url
        record = judgeString("goods", "category", "search", url)  # record momo's type
    elif ("ruten" in url):  # 露天's url
        record = judgeString("item", "category", "find", url)
    if record == 1:
        writeToExcel(1, gender, age, emotion, title, url, IDvalue,time)
    elif record == 2:
        writeToExcel(2, gender, age, emotion, title, url, IDvalue,time)
    elif record == 3:
        writeToExcel(3, gender, age, emotion, title, url, IDvalue,time)


def judgeString(item1, item2, item3, url):
    if item1 in url:
        return 1
    elif item2 in url:
        return 2
    elif item3 in url:
        return 3
    else:
        return 0


def writeToExcel(record, gender, age, emotion, title, url, IDvalue,time):
    if record == 1:
        max_row = s1.max_row
        s1.cell(max_row + 1, 2).value = gender
        s1.cell(max_row + 1, 3).value = age
        s1.cell(max_row + 1, 4).value = emotion
        s1.cell(max_row + 1, 6).value = title
        s1.cell(max_row + 1, 7).value = url
        s1.cell(max_row + 1, 8).value = IDvalue
        s1.cell(max_row + 1, 9).value = time
    elif record == 2:
        max_row = s2.max_row
        print(max_row)
        s2.cell(max_row + 1, 2).value = gender
        s2.cell(max_row + 1, 3).value = age
        s2.cell(max_row + 1, 4).value = emotion
        s2.cell(max_row + 1, 6).value = title
        s2.cell(max_row + 1, 7).value = url
        s2.cell(max_row + 1, 8).value = IDvalue
    elif record == 3:
        max_row = s3.max_row
        s3.cell(max_row + 1, 2).value = gender
        s3.cell(max_row + 1, 3).value = age
        s3.cell(max_row + 1, 4).value = emotion
        s3.cell(max_row + 1, 6).value = title
        s3.cell(max_row + 1, 7).value = url
    wb.save("ComDataSave.xlsx")


def monitor_browser_time(recommand,return_type):
    start_time = time.time()
    while time.time() - start_time <= 1:
        # 檢查是否超過30秒
        time.sleep(1)

    # 超過30秒，顯示特定的GUI
    show_message(recommand,return_type)

import random


# def openURL():
#     global recommand
#     driver2 = webdriver.Chrome()
#     driver2.get(recommand)
# def openURLbyThread():
#     t1 = threading.Thread(target=openURL())
#     t1.start()
def show_message(recommand,return_type):
    # 在GUI中顯示特定訊息
    root = tk.Tk()
    root.title("特定訊息")

    label = tk.Label(root, text='其他人也看了：'+return_type+"---"+recommand)
    label.pack(padx=20, pady=20)
    root.attributes("-topmost", True)
    #tk.Button(root, font=14, text="前往", width=10, height=2, command=lambda: openURLbyThread).pack(side='left',padx=50)

    def show():
        start_time = time.time()
        while time.time() - start_time <= 10:
            time.sleep(1)
        root.destroy()

    t2 = threading.Thread(target=show)
    t2.start()
    root.mainloop()

watch_url_and_title()
